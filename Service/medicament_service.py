from copy import deepcopy

from openpyxl import Workbook

from Domain.add_operation import AddOperation
from Domain.delete_operation import DeleteOperation
from Domain.medicament import Medicament
from Domain.medicament_validator import MedicamentValidator
from Domain.update_operation import UpdateOperation
from Service.undo_redo_service import UndoRedoService


class MedicamentService:
    def __init__(self, medicament_validator: MedicamentValidator, medicament_repository, undo_redo_service: UndoRedoService):
        '''
        Initializeaza un medicament

        :param medicament_validator:
        :param medicament_repository:
        :param undo_redo_service:
        '''
        self.__medicament_repository = medicament_repository
        self.__medicament_validator = medicament_validator
        self.__undo_redo_service = undo_redo_service

    def create(self, id_medicament, nume, producator, pret, necesita_reteta):
        '''
        Creaza un medicament

        :param id_medicament: id-ul medicamentului
        :param nume: numele
        :param producator: producatorul
        :param pret: pretul, float
        :param necesita_reteta: necesita reteta? da sau nu
        :return:
        '''
        medicament = Medicament(id_medicament, nume, producator, pret,  necesita_reteta)
        if medicament.necesita_reteta != 'da' and medicament.necesita_reteta != 'nu':
            raise ValueError('necesita_reteta introdus gresit!')
        self.__medicament_validator.validate(medicament)

        '''
        if necesita_reteta == 'da':
            necesita_reteta = True
        elif necesita_reteta == 'nu':
            necesita_reteta = False
        '''
        self.__medicament_repository.create(medicament)
        self.__undo_redo_service.add_to_undo(AddOperation(self.__medicament_repository, medicament))
        self.__undo_redo_service.clear_redo()

    def delete(self, id_medicament):
        '''
        Sterge un medicament dupa un id

        :param id_medicament: id-ul dupa care se face stergerea
        :return:
        '''
        medicament = self.__medicament_repository.find_by_id(id_medicament)
        self.__medicament_repository.delete(id_medicament)
        self.__undo_redo_service.add_to_undo(DeleteOperation(self.__medicament_repository, medicament))
        self.__undo_redo_service.clear_redo()

    def update(self, id_medicament, nume, producator, pret, necesita_reteta):
        '''
        Modifica un medicament

        :param id_medicament: id-ul medicamentului
        :param nume: numele
        :param producator: producatorul
        :param pret: pretul, float
        :param necesita_reteta: necesita reteta? da sau nu
        :return:
        '''
        medicament1 = self.__medicament_repository.find_by_id(id_medicament)
        medicament2 = self.__medicament_repository.find_by_id(id_medicament)

        if necesita_reteta == 'da':
            necesita_reteta = True
        elif necesita_reteta == 'nu':
            necesita_reteta = False

        if nume != '':
            medicament2.nume = nume
        if producator != '':
            medicament2.producator = producator
        if pret != '':
            medicament2.pret = pret
        if necesita_reteta != '':
            medicament2.necesita_reteta = necesita_reteta

        self.__medicament_repository.update(medicament2)

        self.__undo_redo_service.add_to_undo(UpdateOperation(self.__medicament_repository, medicament1, medicament2))
        self.__undo_redo_service.clear_redo()

    def get_all(self):
        '''
        Returneaza o lista cu toate medicamentele

        :return:
        '''
        return deepcopy(list(self.__medicament_repository.get_all()))


    def export(self):
        '''
        Creaza un fisier excel cu toate medicamenteles

        :return:
        '''
        workbook = Workbook()
        sheet = workbook.active

        sheet["A1"] = "id_medicament"
        sheet["B1"] = "nume"
        sheet["C1"] = "producator"
        sheet["D1"] = "pret"
        sheet["E1"] = "necesita_reteta"

        lst = self.__medicament_repository.get_all()
        for row in range(2, len(lst) + 2):
            sheet.cell(row, 1).value = lst[row - 2].id_entity
            sheet.cell(row, 2).value = lst[row - 2].nume
            sheet.cell(row, 3).value = lst[row - 2].producator
            sheet.cell(row, 4).value = lst[row - 2].pret
            sheet.cell(row, 5).value = lst[row - 2].necesita_reteta

        workbook.save(filename="hello_world.xlsx")

